import os
import re
import fitz  
import time
import json
import pickle
import threading
import logging
from datetime import datetime
from InquirerPy import prompt
from collections import defaultdict
from functools import wraps, lru_cache
from colorama import Fore, Style, init
from openpyxl.styles import PatternFill
from itens import extract_items_from_pdf
from openpyxl import load_workbook, Workbook

init(autoreset=True)


# ==================================================
# CONFIGURAÇÕES GLOBAIS
# ==================================================

REPORTS_FOLDER = os.path.join(os.getcwd(), "relatorios")
VERIFIED_REPORTS_FOLDER = os.path.join(os.getcwd(), "relatorios_verificados")

# pastas reais do CEDUC E DO NEOA
SPREADSHEET_FOLDERS = [
    r"\\fileceduc\grupos\ceduc_secretaria\PATRIMÔNIO\CEDUC_LEVANTAMENTO PATRIMÔNIO_2025",
    r"\\fileceduc\grupos\ceduc_secretaria\PATRIMÔNIO\2025_PATRIMÔNIO_NEOA",
    r"\\fileceduc\grupos\ceduc_secretaria\PATRIMÔNIO\NEI\2024_Inventário"
]

SEARCH_COLUMNS = {
    "1": ("tombamento", 2),
    "2": ("patrimonio", 3),
    "3": ("inventario", 4),
    "4": ("especificacao", 5)
}


# ==================================================
# CLASSE: ERROR HANDLER
# ==================================================

class ErrorHandler:
    """Manipulação de erros e logging"""
    
    def __init__(self):
        self.setup_logging()
    
    def setup_logging(self):
        """Configura sistema de logs"""
        log_dir = os.path.join(os.getcwd(), "logs")
        os.makedirs(log_dir, exist_ok=True)
        
        log_file = os.path.join(log_dir, f"patrimonio_{datetime.now().strftime('%Y%m')}.log")
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def safe_execute(self, func):
        """Decorator para execução segura"""
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                self.logger.error(f"Erro em {func.__name__}: {str(e)}", exc_info=True)
                print(Fore.RED + f"⚠️ Ocorreu um erro em {func.__name__}: {str(e)}")
                print(Fore.YELLOW + "O sistema continuará funcionando...")
                return None
        return wrapper


# ==================================================
# CLASSE: CONFIG MANAGER
# ==================================================

class ConfigManager:
    """Gerenciamento de configurações"""
    
    def __init__(self):
        self.config_file = os.path.join(os.getcwd(), "config.json")
        self.profiles_dir = os.path.join(os.getcwd(), "profiles")
        self.default_config = {
            "interface": {
                "theme": "dark",
                "language": "pt",
                "auto_save": True,
                "notifications": True,
                "show_progress": True
            },
            "performance": {
                "cache_enabled": True,
                "cache_ttl": 3600,
                "batch_size": 50,
                "max_workers": 4
            },
            "paths": {
                "reports_folder": "relatorios",
                "verified_folder": "relatorios_verificados",
                "backup_folder": "backups",
                "exports_folder": "exports",
                "qr_codes_folder": "qr_codes"
            },
            "search": {
                "fuzzy_search": False,
                "partial_match": True,
                "auto_correct": True,
                "case_sensitive": False
            },
            "backup": {
                "auto_backup": True,
                "max_backups": 10,
                "backup_before_changes": True
            }
        }
        self.load_config()
    
    def load_config(self):
        """Carrega ou cria configuração"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    user_config = json.load(f)
                    # Mescla configurações, mantendo padrões para chaves ausentes
                    self.config = self.deep_merge(self.default_config, user_config)
            else:
                self.config = self.default_config.copy()
                self.save_config()
        except Exception as e:
            print(Fore.YELLOW + f"⚠️ Erro ao carregar configuração: {e}")
            self.config = self.default_config.copy()
    
    def save_config(self):
        """Salva configuração"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(Fore.RED + f"❌ Erro ao salvar configuração: {e}")
    
    def deep_merge(self, default, user):
        """Faz merge profundo de dicionários"""
        result = default.copy()
        for key, value in user.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = self.deep_merge(result[key], value)
            else:
                result[key] = value
        return result
    
    def get(self, *keys):
        """Obtém valor de configuração aninhado"""
        value = self.config
        for key in keys:
            value = value.get(key, {})
        return value
    
    def set(self, value, *keys):
        """Define valor de configuração aninhado"""
        config = self.config
        for key in keys[:-1]:
            config = config.setdefault(key, {})
        config[keys[-1]] = value
        self.save_config()
    
    def create_user_profile(self, profile_name):
        """Cria perfil personalizado"""
        os.makedirs(self.profiles_dir, exist_ok=True)
        profile_path = os.path.join(self.profiles_dir, f"{profile_name}.json")
        
        profile = {
            "created": datetime.now().isoformat(),
            "last_used": datetime.now().isoformat(),
            "favorite_searches": [],
            "common_rooms": [],
            "preferences": self.config.copy()
        }
        
        try:
            with open(profile_path, 'w', encoding='utf-8') as f:
                json.dump(profile, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            print(Fore.RED + f"❌ Erro ao criar perfil: {e}")
            return False


# ==================================================
# CLASSE: PERFORMANCE OPTIMIZER
# ==================================================

class PerformanceOptimizer:
    """Otimização de performance com cache"""
    
    def __init__(self, config_manager):
        self.config_manager = config_manager
        self.cache_file = os.path.join(os.getcwd(), ".spreadsheet_cache.pkl")
    
    @lru_cache(maxsize=1)
    def get_spreadsheet_data_cached(self):
        """Cache de dados de planilhas com TTL"""
        if not self.config_manager.get("performance", "cache_enabled"):
            return Model.load_spreadsheet_data()
        
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'rb') as f:
                    cache_data = pickle.load(f)
                    cache_age = time.time() - cache_data['timestamp']
                    
                    if cache_age < self.config_manager.get("performance", "cache_ttl"):
                        print(Fore.CYAN + "📦 Usando dados em cache...")
                        return cache_data['data']
            except Exception as e:
                print(Fore.YELLOW + f"⚠️ Cache corrompido: {e}")
        
        # Recarrega dados
        print(Fore.YELLOW + "🔄 Carregando dados das planilhas...")
        data = Model.load_spreadsheet_data()
        
        # Salva cache
        try:
            with open(self.cache_file, 'wb') as f:
                pickle.dump({'timestamp': time.time(), 'data': data}, f)
        except Exception as e:
            print(Fore.YELLOW + f"⚠️ Não foi possível salvar cache: {e}")
        
        return data
    
    def clear_cache(self):
        """Limpa cache"""
        if os.path.exists(self.cache_file):
            os.remove(self.cache_file)
            print(Fore.GREEN + "✅ Cache limpo")
    
    def batch_process(self, items, callback, batch_size=50):
        """Processa itens em lotes"""
        total = len(items)
        for i in range(0, total, batch_size):
            batch = items[i:i + batch_size]
            callback(batch)
            
            if self.config_manager.get("interface", "show_progress"):
                percent = ((i + len(batch)) / total) * 100
                print(f"\r📊 Progresso: {percent:.1f}%", end="", flush=True)
        
        if self.config_manager.get("interface", "show_progress"):
            print()


# ==================================================
# CLASSE: BACKUP SYSTEM
# ==================================================

class BackupSystem:
    """Sistema de backup e restauração"""
    
    def __init__(self, config_manager):
        self.config_manager = config_manager
        self.backup_dir = os.path.join(os.getcwd(), 
                                      config_manager.get("paths", "backup_folder"))
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def create_backup(self, description="", files_to_backup=None):
        """Cria backup das planilhas"""
        if not self.config_manager.get("backup", "auto_backup"):
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(self.backup_dir, f"backup_{timestamp}")
        os.makedirs(backup_path, exist_ok=True)
        
        metadata = {
            "timestamp": datetime.now().isoformat(),
            "description": description,
            "user": os.getenv('USERNAME', os.getenv('USER', 'Unknown')),
            "files": []
        }
        
        try:
            import shutil
            files_backed_up = 0
            
            for folder in SPREADSHEET_FOLDERS:
                if not os.path.exists(folder):
                    continue
                
                for file in os.listdir(folder):
                    if not file.endswith(".xlsx"):
                        continue
                    
                    if files_to_backup and file not in files_to_backup:
                        continue
                    
                    file_path = os.path.join(folder, file)
                    dest_path = os.path.join(backup_path, file)
                    
                    shutil.copy2(file_path, dest_path)
                    metadata["files"].append({
                        "name": file,
                        "source": folder,
                        "size": os.path.getsize(file_path)
                    })
                    files_backed_up += 1
            
            if files_backed_up > 0:
                with open(os.path.join(backup_path, "metadata.json"), 'w', encoding='utf-8') as f:
                    json.dump(metadata, f, indent=2, ensure_ascii=False)
                
                print(Fore.GREEN + f"✅ Backup criado: {backup_path} ({files_backed_up} arquivos)")
                
                # Limita número de backups
                self.cleanup_old_backups()
                
                return backup_path
            else:
                print(Fore.YELLOW + "⚠️ Nenhum arquivo para backup")
                return None
                
        except Exception as e:
            print(Fore.RED + f"❌ Erro ao criar backup: {e}")
            return None
    
    def cleanup_old_backups(self):
        """Remove backups antigos"""
        max_backups = self.config_manager.get("backup", "max_backups")
        
        try:
            backups = []
            for item in os.listdir(self.backup_dir):
                item_path = os.path.join(self.backup_dir, item)
                if os.path.isdir(item_path) and item.startswith("backup_"):
                    backups.append((item_path, os.path.getctime(item_path)))
            
            backups.sort(key=lambda x: x[1])
            
            while len(backups) > max_backups:
                oldest_backup = backups.pop(0)[0]
                import shutil
                shutil.rmtree(oldest_backup)
                print(Fore.YELLOW + f"🗑️ Backup removido: {os.path.basename(oldest_backup)}")
                
        except Exception as e:
            print(Fore.YELLOW + f"⚠️ Não foi possível limpar backups antigos: {e}")
    
    def list_backups(self):
        """Lista backups disponíveis"""
        backups = []
        for item in os.listdir(self.backup_dir):
            item_path = os.path.join(self.backup_dir, item)
            if os.path.isdir(item_path) and item.startswith("backup_"):
                metadata_file = os.path.join(item_path, "metadata.json")
                if os.path.exists(metadata_file):
                    try:
                        with open(metadata_file, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                        backups.append({
                            "path": item_path,
                            "name": item,
                            "timestamp": metadata.get("timestamp", ""),
                            "description": metadata.get("description", ""),
                            "files": len(metadata.get("files", []))
                        })
                    except:
                        backups.append({"path": item_path, "name": item})
        
        return sorted(backups, key=lambda x: x.get("timestamp", ""), reverse=True)
    
    def restore_backup(self, backup_path):
        """Restaura backup"""
        metadata_file = os.path.join(backup_path, "metadata.json")
        
        if not os.path.exists(metadata_file):
            print(Fore.RED + "❌ Backup inválido: metadata.json não encontrado")
            return False
        
        try:
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
            
            print(Fore.CYAN + f"📋 Backup de: {metadata.get('timestamp', 'Desconhecido')}")
            print(Fore.CYAN + f"📝 Descrição: {metadata.get('description', 'Nenhuma')}")
            print(Fore.CYAN + f"👤 Usuário: {metadata.get('user', 'Desconhecido')}")
            print(Fore.CYAN + f"📄 Arquivos: {len(metadata.get('files', []))}")
            
            if not EnhancedView.confirm_dialog("Deseja restaurar este backup?", dangerous=True):
                return False
            
            import shutil
            restored_files = 0
            
            for file_info in metadata.get("files", []):
                file_name = file_info.get("name")
                source_folder = file_info.get("source")
                
                if not file_name or not source_folder:
                    continue
                
                src = os.path.join(backup_path, file_name)
                dest = os.path.join(source_folder, file_name)
                
                if os.path.exists(src):
                    # Cria backup do arquivo atual antes de substituir
                    current_backup = os.path.join(backup_path, f"original_{file_name}")
                    if os.path.exists(dest):
                        shutil.copy2(dest, current_backup)
                    
                    shutil.copy2(src, dest)
                    restored_files += 1
                    print(Fore.GREEN + f"  ✅ Restaurado: {file_name}")
            
            print(Fore.GREEN + f"✅ Restauração concluída: {restored_files} arquivos")
            return True
            
        except Exception as e:
            print(Fore.RED + f"❌ Erro ao restaurar backup: {e}")
            return False


# ==================================================
# CLASSE: ENHANCED VIEW
# ==================================================

class EnhancedView:
    """View aprimorada com mais recursos visuais"""
    
    @staticmethod
    def show_message(message_type, message, details=""):
        """Mensagens padronizadas com ícones"""
        icons = {
            "success": "✅",
            "error": "❌",
            "warning": "⚠️",
            "info": "ℹ️",
            "loading": "🔄",
            "search": "🔎",
            "file": "📄",
            "folder": "📂",
            "stats": "📊",
            "backup": "💾"
        }
        
        colors = {
            "success": Fore.GREEN,
            "error": Fore.RED,
            "warning": Fore.YELLOW,
            "info": Fore.CYAN,
            "loading": Fore.MAGENTA,
            "search": Fore.BLUE,
            "file": Fore.MAGENTA,
            "folder": Fore.BLUE,
            "stats": Fore.CYAN,
            "backup": Fore.YELLOW
        }
        
        icon = icons.get(message_type, "•")
        color = colors.get(message_type, Fore.WHITE)
        
        if details:
            print(f"{color}{icon} {message}: {details}{Style.RESET_ALL}")
        else:
            print(f"{color}{icon} {message}{Style.RESET_ALL}")
    
    @staticmethod
    def progress_bar(iteration, total, prefix='', suffix='', length=50, fill='█'):
        """Barra de progresso visual"""
        percent = f"{100 * (iteration / float(total)):.1f}"
        filled_length = int(length * iteration // total)
        bar = fill * filled_length + '-' * (length - filled_length)
        
        # Cores baseadas no progresso
        if percent >= 75:
            color = Fore.GREEN
        elif percent >= 50:
            color = Fore.YELLOW
        else:
            color = Fore.RED
        
        print(f'\r{color}{prefix} |{bar}| {percent}% {suffix}{Style.RESET_ALL}', end='\r')
        if iteration == total: 
            print()
    
    @staticmethod
    def clear_terminal():
        """Limpa o terminal"""
        os.system("cls" if os.name == "nt" else "clear")
    
    @staticmethod
    def print_line():
        """Imprime linha decorativa"""
        print(Fore.CYAN + "." * 70)
    
    @staticmethod
    def print_header(title, subtitle=""):
        """Imprime cabeçalho formatado"""
        EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()
        print(Fore.YELLOW + Style.BRIGHT + title.center(70))
        if subtitle:
            print(Fore.CYAN + subtitle.center(70))
        EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()

    @staticmethod
    def highlight_key(text, key, key_color=Fore.GREEN):
        """Destaca uma palavra-chave no texto"""
        return text.replace(
            key,
            key_color + Style.BRIGHT + key + Style.RESET_ALL + Fore.WHITE
        )
    
    @staticmethod
    def confirm_dialog(message, dangerous=False):
        """Diálogo de confirmação aprimorado"""
        if dangerous:
            print(Fore.RED + Style.BRIGHT + "⚠️ PERIGO: " + message)
        else:
            print(Fore.YELLOW + message)
        
        response = input(Fore.CYAN + "Confirma? (sim/não): ").lower().strip()
        return response in ['s', 'sim', 'y', 'yes']
    
    @staticmethod
    def tui_enhanced_menu():
        """Menu principal aprimorado"""
        print(Fore.YELLOW + Style.BRIGHT + "OPÇÕES DISPONÍVEIS:")
        print(Fore.CYAN + "Ctrl + C para voltar / sair\n")
        
        pergunta = [{
            "type": "list",
            "name": "opcao",
            "message": "",
            "choices": [
                {"name": "🔍 Buscar por Número de Tombamento", "value": "1"},
                {"name": "🔍 Buscar por Número de Patrimônio", "value": "2"},
                {"name": "🔍 Buscar por Número de Inventário", "value": "3"},
                {"name": "🔍 Buscar por Especificação", "value": "4"},
                {"name": "📄 Verificar itens do relatório PDF", "value": "5"},
                {"name": "📊 Gerar estatísticas", "value": "6"},
                {"name": "💾 Criar backup", "value": "7"},
                {"name": "🔄 Restaurar backup", "value": "8"},
                {"name": "⚙️ Configurações", "value": "9"},
                {"name": "📤 Exportar dados", "value": "10"},
                {"name": "🆚 Comparar relatórios", "value": "11"},
                {"name": "🧹 Limpar cache", "value": "12"},
                {"name": "❌ Sair", "value": "0"},
            ],
        }]
        
        try:
            resposta = prompt(pergunta)
            return resposta["opcao"]
        except KeyboardInterrupt:
            return None
    
    @staticmethod
    def tui_select_pdf(pdf_files):
        """Seleção de PDFs"""
        print(Fore.YELLOW + Style.BRIGHT + "RELATÓRIOS DISPONÍVEIS:")
        print(Fore.CYAN + "Ctrl + C para voltar\n")
        
        pergunta = [{
            "type": "list",
            "name": "pdf",
            "message": "",
            "choices": pdf_files,
        }]
        
        try:
            resposta = prompt(pergunta)
            return resposta["pdf"]
        except KeyboardInterrupt:
            return None
    
    @staticmethod
    def tui_select_backup(backups):
        """Seleção de backup"""
        if not backups:
            print(Fore.YELLOW + "Nenhum backup disponível")
            return None
        
        choices = []
        for i, backup in enumerate(backups, 1):
            timestamp = backup.get("timestamp", "Desconhecido")
            if timestamp != "Desconhecido":
                try:
                    dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    timestamp = dt.strftime("%d/%m/%Y %H:%M")
                except:
                    pass
            
            desc = backup.get("description", "")
            files = backup.get("files", 0)
            
            display = f"{timestamp} | {desc[:30]}{'...' if len(desc) > 30 else ''} | {files} arquivos"
            choices.append({"name": display, "value": backup["path"]})
        
        pergunta = [{
            "type": "list",
            "name": "backup",
            "message": "Selecione o backup para restaurar:",
            "choices": choices,
        }]
        
        try:
            resposta = prompt(pergunta)
            return resposta["backup"]
        except KeyboardInterrupt:
            return None
    
    @staticmethod
    def display_statistics(stats):
        """Exibe estatísticas formatadas"""
        print(Fore.CYAN + Style.BRIGHT + "\n📊 ESTATÍSTICAS DO RELATÓRIO")
        EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()
        print(f"📈 Total de itens: {stats['total']}")
        print(f"✅ Encontrados: {stats['found']} ({stats['percentage_found']:.1f}%)")
        print(f"❌ Não encontrados: {stats['not_found']} ({stats['percentage_not_found']:.1f}%)")
        
        if stats.get('rooms'):
            print("\n📍 Distribuição por sala:")
            for sala, count in stats['rooms'].items():
                print(f"  {sala}: {count} itens")
        
        EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()
    
    @staticmethod
    def display_search_results(item_data, sala, criterion, original_value):
        """Exibe resultados da busca"""
        if criterion == "especificacao":
            print(Fore.MAGENTA + Style.BRIGHT + 
                  f"\n🔎 Procurando por: {Fore.YELLOW}{original_value}\n")
        else:
            print(Fore.MAGENTA + Style.BRIGHT + 
                  f"\n🔎 Procurando pelo item com número de "
                  f"{Fore.YELLOW}{criterion.upper()}: {original_value}\n")
        
        print(Fore.YELLOW + Style.BRIGHT + f"📂 Vasculhando pasta: {item_data['folder']}")
        
        if not os.path.exists(item_data['folder']):
            print(Fore.RED + "  Pasta não encontrada.\n")
            return
        
        print("\n")
        print(Fore.WHITE + "Origem: " + item_data['origin'])
        print(Fore.WHITE + f"Sala (arquivo): {item_data['file']}")
        print(Fore.WHITE + f"Aba: {item_data['sheet']}")
        print(Fore.WHITE + f"Linha: {item_data['row']}")
        
        EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()
        print(Fore.CYAN + f"Item: {item_data['item']}")
        print(Fore.CYAN + f"Tombamento: {item_data['tombamento']}")
        print(Fore.CYAN + f"Patrimônio: {item_data['patrimonio']}")
        print(Fore.CYAN + f"Inventário: {item_data['inventario']}")
        print(Fore.CYAN + f"Especificação: {item_data['especificacao']}")
        print(Fore.CYAN + f"TR: {item_data['tr']}")
        print(Fore.CYAN + f"Situação: {item_data['situacao']}")
        EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()
    
    @staticmethod
    def display_report_results(itens_por_sala):
        """Exibe resultados da verificação de relatório"""
        for sala, itens in itens_por_sala.items():
            print(Fore.CYAN + Style.BRIGHT + f"\n📍 ITENS DA SALA - {sala}")
            EnhancedView.print_line()  # Corrigido: chama EnhancedView.print_line()
            
            for item in itens:
                if item["status"]:
                    print(Fore.GREEN + Style.BRIGHT +
                          f"✔ Tombamento: {item['tombamento']} | Item: {item['denominacao']}")
                else:
                    print(Fore.RED + Style.BRIGHT +
                          f"✖ Tombamento: {item['tombamento']} | Item: {item['denominacao']}")

# ==================================================
# CLASSE: ENHANCED FEATURES
# ==================================================

class EnhancedFeatures:
    """Funcionalidades adicionais"""
    
    def __init__(self, config_manager):
        self.config_manager = config_manager
    
    def generate_statistics(self, tombamento_results, itens_por_sala=None):
        """Gera estatísticas detalhadas"""
        total = len(tombamento_results)
        found = sum(tombamento_results.values())
        not_found = total - found
        
        stats = {
            "total": total,
            "found": found,
            "not_found": not_found,
            "percentage_found": (found / total * 100) if total > 0 else 0,
            "percentage_not_found": (not_found / total * 100) if total > 0 else 0,
            "timestamp": datetime.now().isoformat()
        }
        
        if itens_por_sala:
            stats['rooms'] = {sala: len(itens) for sala, itens in itens_por_sala.items()}
        
        return stats
    
    def export_results(self, data, format_type="excel", filename_prefix="resultados"):
        """Exporta resultados em múltiplos formatos"""
        export_dir = os.path.join(os.getcwd(), self.config_manager.get("paths", "exports_folder"))
        os.makedirs(export_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format_type == "excel":
            filename = os.path.join(export_dir, f"{filename_prefix}_{timestamp}.xlsx")
            self._export_to_excel(data, filename)
            
        elif format_type == "csv":
            filename = os.path.join(export_dir, f"{filename_prefix}_{timestamp}.csv")
            self._export_to_csv(data, filename)
            
        elif format_type == "json":
            filename = os.path.join(export_dir, f"{filename_prefix}_{timestamp}.json")
            self._export_to_json(data, filename)
        
        else:
            print(Fore.RED + f"❌ Formato não suportado: {format_type}")
            return None
        
        EnhancedView.show_message("success", f"Resultados exportados", filename)
        return filename
    
    def _export_to_excel(self, data, filename):
        """Exporta para Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        # Cabeçalhos com formatação
        headers = ["Tombamento", "Status", "Sala", "Denominação", "Data Verificação", "Observações"]
        ws.append(headers)
        
        # Adiciona dados
        for item in data:
            status = "ENCONTRADO" if item.get('status') else "NÃO ENCONTRADO"
            ws.append([
                item.get('tombamento', ''),
                status,
                item.get('sala', ''),
                item.get('denominacao', ''),
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                item.get('observacoes', '')
            ])
        
        wb.save(filename)
    
    def _export_to_csv(self, data, filename):
        """Exporta para CSV"""
        import csv
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(["Tombamento", "Status", "Sala", "Denominação", "Data Verificação"])
            for item in data:
                status = "ENCONTRADO" if item.get('status') else "NÃO ENCONTRADO"
                writer.writerow([
                    item.get('tombamento', ''),
                    status,
                    item.get('sala', ''),
                    item.get('denominacao', ''),
                    datetime.now().strftime("%d/%m/%Y %H:%M")
                ])
    
    def _export_to_json(self, data, filename):
        """Exporta para JSON"""
        export_data = {
            "metadata": {
                "export_date": datetime.now().isoformat(),
                "total_items": len(data),
                "found_items": sum(1 for item in data if item.get('status')),
                "not_found_items": sum(1 for item in data if not item.get('status'))
            },
            "items": data
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
    
    def compare_reports(self, report1_path, report2_path):
        """Compara dois relatórios e mostra diferenças"""
        try:
            items1 = extract_items_from_pdf(report1_path)
            items2 = extract_items_from_pdf(report2_path)
            
            tombamentos1 = {item['tombamento'] for item in items1}
            tombamentos2 = {item['tombamento'] for item in items2}
            
            only_in_1 = tombamentos1 - tombamentos2
            only_in_2 = tombamentos2 - tombamentos1
            common = tombamentos1 & tombamentos2
            
            print(Fore.CYAN + Style.BRIGHT + "\n📊 COMPARAÇÃO DE RELATÓRIOS")
            EnhancedView.print_line()
            print(f"📄 Relatório 1: {os.path.basename(report1_path)} ({len(items1)} itens)")
            print(f"📄 Relatório 2: {os.path.basename(report2_path)} ({len(items2)} itens)")
            EnhancedView.print_line()
            print(f"✅ Itens em comum: {len(common)}")
            print(f"➕ Exclusivos no Relatório 1: {len(only_in_1)}")
            print(f"➕ Exclusivos no Relatório 2: {len(only_in_2)}")
            
            if only_in_1:
                print(Fore.YELLOW + "\n📋 Itens apenas no Relatório 1:")
                for tombamento in sorted(only_in_1)[:10]:  # Mostra apenas os 10 primeiros
                    item = next((i for i in items1 if i['tombamento'] == tombamento), None)
                    if item:
                        print(f"  • {tombamento}: {item.get('denominacao', '')}")
                if len(only_in_1) > 10:
                    print(f"  ... e mais {len(only_in_1) - 10} itens")
            
            return {
                "only_in_1": only_in_1,
                "only_in_2": only_in_2,
                "common": common,
                "report1_count": len(items1),
                "report2_count": len(items2)
            }
            
        except Exception as e:
            print(Fore.RED + f"❌ Erro ao comparar relatórios: {e}")
            return None
    
    def generate_qr_code(self, tombamento, item_info):
        """Gera QR Code para tombamento (opcional)"""
        try:
            import qrcode
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            
            data = {
                "tombamento": tombamento,
                "item": item_info.get('denominacao', ''),
                "sala": item_info.get('sala', ''),
                "data_verificacao": datetime.now().isoformat(),
                "status": item_info.get('status', False)
            }
            
            qr.add_data(json.dumps(data, ensure_ascii=False))
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            
            qr_dir = os.path.join(os.getcwd(), self.config_manager.get("paths", "qr_codes_folder"))
            os.makedirs(qr_dir, exist_ok=True)
            
            filename = os.path.join(qr_dir, f"qr_{tombamento}.png")
            img.save(filename)
            
            EnhancedView.show_message("success", f"QR Code gerado", filename)
            return filename
            
        except ImportError:
            print(Fore.YELLOW + "⚠️ Biblioteca qrcode não instalada. Use: pip install qrcode[pil]")
            return None
        except Exception as e:
            print(Fore.YELLOW + f"⚠️ Erro ao gerar QR Code: {e}")
            return None


# ==================================================
# CLASSE: MODEL
# ==================================================

class Model:
    """Responsável por manipulação de dados e arquivos"""
    
    @staticmethod
    def extract_room_from_filename(filename):
        """Extrai código + nome da sala a partir do nome do arquivo"""
        name = os.path.splitext(filename)[0]
        name = re.sub(r"\(.*?\)", "", name)
        name = name.strip()
        parts = [p.strip() for p in name.split("_") if p.strip()]
        
        if not parts:
            return name
        
        room_parts = []
        
        for i, part in enumerate(parts):
            upper = part.upper()
            
            if i > 0 and any(keyword in upper for keyword in [
                "CEDUC", "NEOA", "SUPORTE", "GABINETE", "PROF", "DOCENTE",
                "AL", "ANA", "CAROLINA", "RODRIGUES", "OLIVEIRA"
            ]):
                break
            
            room_parts.append(part)
        
        return " ".join(room_parts)
    
    @staticmethod
    def list_pdf_reports():
        """Lista arquivos PDF na pasta de relatórios"""
        if not os.path.exists(REPORTS_FOLDER):
            EnhancedView.show_message("error", "Pasta 'relatorios' não encontrada")
            return []
        
        pdf_files = [
            f for f in os.listdir(REPORTS_FOLDER)
            if f.lower().endswith(".pdf")
        ]
        
        if not pdf_files:
            EnhancedView.show_message("warning", "Nenhum relatório PDF encontrado")
            return []
        
        return pdf_files
    
    @staticmethod
    def load_spreadsheet_data():
        """Carrega todos os tombamentos das planilhas"""
        found_tombamentos = {}
        
        for folder in SPREADSHEET_FOLDERS:
            if not os.path.exists(folder):
                EnhancedView.show_message("warning", f"Pasta não encontrada", folder)
                continue
            
            for file in os.listdir(folder):
                if not file.endswith(".xlsx"):
                    continue
                
                file_path = os.path.join(folder, file)
                
                try:
                    workbook = load_workbook(file_path, data_only=True)
                except Exception as e:
                    EnhancedView.show_message("warning", f"Erro ao abrir arquivo", file)
                    continue
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 3:
                            continue
                        
                        tombamento_cell = str(row[2]).strip() if row[2] else ""
                        if tombamento_cell and tombamento_cell not in found_tombamentos:
                            sala = Model.extract_room_from_filename(file)
                            found_tombamentos[tombamento_cell] = sala
        
        EnhancedView.show_message("success", f"Carregados {len(found_tombamentos)} tombamentos")
        return found_tombamentos
    
    @staticmethod
    def draw_check(page, x, y, size, color, width=1.5):
        """Desenha um ✓ vetorial usando 2 linhas"""
        page.draw_line(
            fitz.Point(x - size * 0.4, y),
            fitz.Point(x - size * 0.1, y + size * 0.4),
            color=color,
            width=width
        )
        
        page.draw_line(
            fitz.Point(x - size * 0.1, y + size * 0.4),
            fitz.Point(x + size * 0.5, y - size * 0.5),
            color=color,
            width=width
        )
    
    @staticmethod
    def draw_x(page, x, y, size, color, width=1.5):
        """Desenha um ✗ vetorial usando 2 linhas"""
        page.draw_line(
            fitz.Point(x - size * 0.5, y - size * 0.5),
            fitz.Point(x + size * 0.5, y + size * 0.5),
            color=color,
            width=width
        )
        
        page.draw_line(
            fitz.Point(x - size * 0.5, y + size * 0.5),
            fitz.Point(x + size * 0.5, y - size * 0.5),
            color=color,
            width=width
        )
    
    @staticmethod
    def generate_checked_pdf(original_pdf, output_pdf, tombamento_results):
        """Gera uma cópia do PDF com ícones vetoriais"""
        doc = fitz.open(original_pdf)
        
        GREEN = (0.0, 0.6, 0.25)
        RED = (0.75, 0.15, 0.15)
        
        ICON_X = 16      # distância da borda esquerda
        ICON_SIZE = 8    # tamanho do ícone
        STROKE = 1.5     # espessura do traço
        
        for page in doc:
            words = page.get_text("words")
            
            for tombamento, found in tombamento_results.items():
                line_words = [w for w in words if tombamento in w[4]]
                if not line_words:
                    continue
                
                for w in line_words:
                    _, y0, _, y1, _, block, line_no, _ = w
                    
                    same_line = [
                        word for word in words
                        if word[5] == block and word[6] == line_no
                    ]
                    if not same_line:
                        continue
                    
                    center_y = (y0 + y1) / 2
                    color = GREEN if found else RED
                    
                    if found:
                        Model.draw_check(page, ICON_X, center_y, ICON_SIZE, color, STROKE)
                    else:
                        Model.draw_x(page, ICON_X, center_y, ICON_SIZE, color, STROKE)
        
        doc.save(output_pdf)
        doc.close()


# ==================================================
# CLASSE: ENHANCED CONTROLLER
# ==================================================

class EnhancedController:
    """Controller aprimorado com todas as funcionalidades"""
    
    def __init__(self):
        self.view = EnhancedView()
        self.model = Model()
        
        # Sistemas auxiliares
        self.error_handler = ErrorHandler()
        self.config_manager = ConfigManager()
        self.performance_optimizer = PerformanceOptimizer(self.config_manager)
        self.backup_system = BackupSystem(self.config_manager)
        self.enhanced_features = EnhancedFeatures(self.config_manager)
        
        # Configurações
        self.GOOD_FILL = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
        self.BAD_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
        
        # Estado da aplicação
        self.current_session = {
            "start_time": datetime.now(),
            "searches_count": 0,
            "reports_processed": 0,
            "backups_created": 0
        }
    
    def search_items(self, column_index, value, criterion):
        """Busca itens nas planilhas (com cache)"""
        original_value = value.strip()
        search_value = original_value.upper() if not self.config_manager.get("search", "case_sensitive") else original_value
        criterion = criterion.lower()
        
        EnhancedView.show_message("search", f"Buscando por {criterion}", original_value)
        
        # Usa cache se habilitado
        if self.config_manager.get("performance", "cache_enabled"):
            spreadsheet_data = self.performance_optimizer.get_spreadsheet_data_cached()
            # Converte para formato de busca
            found_tombamentos = spreadsheet_data
        else:
            found_tombamentos = self.model.load_spreadsheet_data()
        
        found = False
        
        for folder in SPREADSHEET_FOLDERS:
            if not os.path.exists(folder):
                continue
            
            EnhancedView.show_message("folder", "Vasculhando pasta", folder)
            
            for file in os.listdir(folder):
                if not file.endswith(".xlsx"):
                    continue
                
                file_path = os.path.join(folder, file)
                
                try:
                    workbook = load_workbook(file_path, data_only=True)
                except:
                    continue
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or len(row) < 8:
                            continue
                        
                        cell_value = str(row[column_index]).strip() if row[column_index] else ""
                        
                        # Aplica configurações de busca
                        if not self.config_manager.get("search", "case_sensitive"):
                            cell_value = cell_value.upper()
                        
                        match_found = False
                        if self.config_manager.get("search", "partial_match"):
                            match_found = search_value in cell_value
                        else:
                            match_found = cell_value == search_value
                        
                        if match_found:
                            found = True
                            self._display_search_result(row, row_index, file, sheet_name, folder)
        
        self.current_session["searches_count"] += 1
        
        if not found:
            EnhancedView.show_message("error", "Nenhum resultado encontrado")
    
    def _display_search_result(self, row, row_index, file, sheet_name, folder):
        """Exibe resultado da busca"""
        # Identificação da origem
        if "NEOA" in folder.upper():
            origin = Fore.GREEN + Style.BRIGHT + "NEOA"
        elif "CEDUC" in folder.upper():
            origin = Fore.BLUE + Style.BRIGHT + "CEDUC"
        else:
            origin = Fore.WHITE + "DESCONHECIDA"
        
        print("\n")
        print(Fore.WHITE + "Origem: " + origin)
        print(Fore.WHITE + f"Sala (arquivo): {file}")
        print(Fore.WHITE + f"Aba: {sheet_name}")
        print(Fore.WHITE + f"Linha: {row_index}")
        
        EnhancedView.print_line()
        print(Fore.CYAN + f"Item: {row[1]}")
        print(Fore.CYAN + f"Tombamento: {row[2]}")
        print(Fore.CYAN + f"Patrimônio: {row[3]}")
        print(Fore.CYAN + f"Inventário: {row[4]}")
        print(Fore.CYAN + f"Especificação: {row[5]}")
        print(Fore.CYAN + f"TR: {row[6]}")
        print(Fore.CYAN + f"Situação: {row[7]}")
        EnhancedView.print_line()
    
    @ErrorHandler().safe_execute
    def search_items_from_pdf(self, pdf_path):
        """Processa itens de um relatório PDF (com cache)"""
        os.makedirs(VERIFIED_REPORTS_FOLDER, exist_ok=True)
        
        EnhancedView.show_message("loading", "Extraindo itens do PDF")
        
        loader = threading.Thread(
            target=EnhancedView.show_message,
            args=("loading", "Processando...")
        )
        loader.start()
        
        report_items = extract_items_from_pdf(pdf_path)
        loader.join()
        
        if not report_items:
            EnhancedView.show_message("error", "Nenhum item encontrado no relatório")
            return
        
        EnhancedView.show_message("info", f"Encontrados {len(report_items)} itens no relatório")
        
        # Usa cache para dados das planilhas
        found_tombamentos = self.performance_optimizer.get_spreadsheet_data_cached()
        
        tombamento_results = {}
        itens_por_sala = defaultdict(list)
        export_data = []
        
        # Processamento em lotes
        batch_size = self.config_manager.get("performance", "batch_size")
        
        def process_batch(batch):
            for item in batch:
                tombamento = item["tombamento"]
                denominacao = item["denominacao"]
                
                if tombamento in found_tombamentos:
                    sala = found_tombamentos[tombamento]
                    tombamento_results[tombamento] = True
                    
                    itens_por_sala[sala].append({
                        "status": True,
                        "tombamento": tombamento,
                        "denominacao": denominacao,
                        "sala": sala
                    })
                    
                    export_data.append({
                        "tombamento": tombamento,
                        "status": True,
                        "sala": sala,
                        "denominacao": denominacao
                    })
                else:
                    tombamento_results[tombamento] = False
                    
                    itens_por_sala["NÃO ENCONTRADO"].append({
                        "status": False,
                        "tombamento": tombamento,
                        "denominacao": denominacao,
                        "sala": "NÃO ENCONTRADO"
                    })
                    
                    export_data.append({
                        "tombamento": tombamento,
                        "status": False,
                        "sala": "NÃO ENCONTRADO",
                        "denominacao": denominacao
                    })
        
        # Executa processamento em lotes
        self.performance_optimizer.batch_process(report_items, process_batch, batch_size)
        
        # Exibe resultados
        EnhancedView.show_message("success", "Verificação concluída")
        
        for sala, itens in itens_por_sala.items():
            print(Fore.CYAN + Style.BRIGHT + f"\n📍 ITENS DA SALA - {sala}")
            EnhancedView.print_line()
            
            for item in itens:
                if item["status"]:
                    print(Fore.GREEN + Style.BRIGHT +
                          f"✔ Tombamento: {item['tombamento']} | Item: {item['denominacao']}")
                else:
                    print(Fore.RED + Style.BRIGHT +
                          f"✖ Tombamento: {item['tombamento']} | Item: {item['denominacao']}")
        
        # Gera estatísticas
        stats = self.enhanced_features.generate_statistics(tombamento_results, itens_por_sala)
        EnhancedView.display_statistics(stats)
        
        # Pergunta sobre aplicação nas planilhas
        if EnhancedView.confirm_dialog("Deseja aplicar o resultado nas planilhas?"):
            if self.config_manager.get("backup", "backup_before_changes"):
                self.backup_system.create_backup("Backup antes de aplicar alterações")
            
            self.apply_results_to_spreadsheets(tombamento_results)
        
        # Pergunta sobre exportação
        if EnhancedView.confirm_dialog("Deseja exportar os resultados?"):
            format_choice = input("Formato (excel/csv/json): ").lower().strip()
            if format_choice in ["excel", "csv", "json"]:
                self.enhanced_features.export_results(export_data, format_choice)
        
        # Gera PDF marcado
        original_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_filename = f"{original_name} - verificado.pdf"
        output_pdf = os.path.join(VERIFIED_REPORTS_FOLDER, output_filename)
        
        if os.path.exists(output_pdf):
            EnhancedView.show_message("warning", "PDF já verificado anteriormente", output_pdf)
        else:
            self.model.generate_checked_pdf(pdf_path, output_pdf, tombamento_results)
            EnhancedView.show_message("success", "PDF gerado com marcações", output_pdf)
        
        self.current_session["reports_processed"] += 1
    
    def apply_results_to_spreadsheets(self, tombamento_results):
        """Aplica resultados às planilhas"""
        EnhancedView.show_message("warning", "MODO PRODUÇÃO", "Alterações não podem ser desfeitas automaticamente")
        
        if not EnhancedView.confirm_dialog("Você está prestes a alterar TODAS as planilhas.", dangerous=True):
            EnhancedView.show_message("info", "Operação cancelada")
            return
        
        total_files = 0
        total_changes = 0
        modified_files = []
        
        for folder in SPREADSHEET_FOLDERS:
            if not os.path.exists(folder):
                continue
            
            for file in os.listdir(folder):
                if not file.endswith(".xlsx"):
                    continue
                
                file_path = os.path.join(folder, file)
                
                try:
                    wb = load_workbook(file_path)
                except Exception:
                    continue
                
                file_changes = 0
                
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    
                    for row in ws.iter_rows(min_row=2):
                        tombamento_cell = row[2]
                        
                        if tombamento_cell.value is None:
                            continue
                        
                        tombamento = str(tombamento_cell.value).strip()
                        
                        if tombamento in tombamento_results:
                            status = tombamento_results[tombamento]
                            target_cell = row[0]
                            
                            target_cell.fill = self.GOOD_FILL if status else self.BAD_FILL
                            file_changes += 1
                
                if file_changes > 0:
                    wb.save(file_path)
                    total_files += 1
                    total_changes += file_changes
                    modified_files.append(file)
                    
                    EnhancedView.show_message("success", f"Arquivo atualizado", f"{file} ({file_changes} linhas)")
        
        EnhancedView.show_message("success", 
            f"PROCESSO FINALIZADO: {total_files} planilhas alteradas, {total_changes} linhas atualizadas")
        
        # Cria backup das alterações
        if modified_files and self.config_manager.get("backup", "auto_backup"):
            self.backup_system.create_backup("Backup após aplicação de resultados", modified_files)
    
    def handle_statistics(self):
        """Manipula geração de estatísticas"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("GERAR ESTATÍSTICAS")
        
        # Carrega dados para estatísticas
        data = self.performance_optimizer.get_spreadsheet_data_cached()
        
        if not data:
            EnhancedView.show_message("error", "Nenhum dado disponível")
            return
        
        stats = {
            "total_tombamentos": len(data),
            "salas_unicas": len(set(data.values())),
            "tombamentos_por_sala": {}
        }
        
        # Contagem por sala
        for tombamento, sala in data.items():
            stats["tombamentos_por_sala"][sala] = stats["tombamentos_por_sala"].get(sala, 0) + 1
        
        print(Fore.CYAN + Style.BRIGHT + "\n📊 ESTATÍSTICAS GERAIS")
        EnhancedView.print_line()
        print(f"📈 Total de tombamentos: {stats['total_tombamentos']}")
        print(f"📍 Salas únicas: {stats['salas_unicas']}")
        
        print("\n📋 Tombamentos por sala (top 10):")
        sorted_salas = sorted(stats["tombamentos_por_sala"].items(), key=lambda x: x[1], reverse=True)[:10]
        
        for i, (sala, count) in enumerate(sorted_salas, 1):
            print(f"  {i}. {sala}: {count} tombamentos")
        
        EnhancedView.print_line()
        
        # Pergunta sobre exportação
        if EnhancedView.confirm_dialog("Deseja exportar as estatísticas?"):
            export_data = [{"sala": sala, "tombamentos": count} for sala, count in stats["tombamentos_por_sala"].items()]
            self.enhanced_features.export_results(export_data, "json", "estatisticas")
    
    def handle_backup(self):
        """Manipula criação de backup"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("CRIAR BACKUP")
        
        description = input(Fore.YELLOW + "Descrição do backup (opcional): ").strip()
        
        EnhancedView.show_message("loading", "Criando backup")
        backup_path = self.backup_system.create_backup(description)
        
        if backup_path:
            EnhancedView.show_message("success", "Backup criado com sucesso")
        else:
            EnhancedView.show_message("error", "Falha ao criar backup")
        
        self.current_session["backups_created"] += 1
    
    def handle_restore(self):
        """Manipula restauração de backup"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("RESTAURAR BACKUP")
        
        backups = self.backup_system.list_backups()
        
        if not backups:
            EnhancedView.show_message("warning", "Nenhum backup disponível")
            return
        
        selected_backup = EnhancedView.tui_select_backup(backups)
        
        if selected_backup:
            success = self.backup_system.restore_backup(selected_backup)
            if success:
                EnhancedView.show_message("success", "Backup restaurado com sucesso")
    
    def handle_configuration(self):
        """Manipula configurações do sistema"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("CONFIGURAÇÕES")
        
        print(Fore.CYAN + "\n⚙️ Configurações atuais:")
        EnhancedView.print_line()
        
        # Exibe configurações principais
        config_display = {
            "Cache habilitado": self.config_manager.get("performance", "cache_enabled"),
            "Tamanho do lote": self.config_manager.get("performance", "batch_size"),
            "Backup automático": self.config_manager.get("backup", "auto_backup"),
            "Busca parcial": self.config_manager.get("search", "partial_match"),
            "Case sensitive": self.config_manager.get("search", "case_sensitive")
        }
        
        for key, value in config_display.items():
            status = Fore.GREEN + "✅" if value else Fore.RED + "❌"
            print(f"{status} {key}: {value}")
        
        EnhancedView.print_line()
        
        # Opções de configuração
        print(Fore.YELLOW + "\nOpções:")
        print("1. Alternar cache")
        print("2. Alterar tamanho do lote")
        print("3. Alternar backup automático")
        print("4. Voltar")
        
        choice = input(Fore.CYAN + "\nEscolha uma opção: ").strip()
        
        if choice == "1":
            current = self.config_manager.get("performance", "cache_enabled")
            self.config_manager.set(not current, "performance", "cache_enabled")
            EnhancedView.show_message("success", f"Cache {'habilitado' if not current else 'desabilitado'}")
            
        elif choice == "2":
            try:
                new_size = int(input("Novo tamanho do lote (10-1000): "))
                if 10 <= new_size <= 1000:
                    self.config_manager.set(new_size, "performance", "batch_size")
                    EnhancedView.show_message("success", f"Tamanho do lote alterado para {new_size}")
                else:
                    EnhancedView.show_message("error", "Valor fora do intervalo permitido")
            except ValueError:
                EnhancedView.show_message("error", "Valor inválido")
        
        elif choice == "3":
            current = self.config_manager.get("backup", "auto_backup")
            self.config_manager.set(not current, "backup", "auto_backup")
            EnhancedView.show_message("success", f"Backup automático {'habilitado' if not current else 'desabilitado'}")
    
    def handle_export(self):
        """Manipula exportação de dados"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("EXPORTAR DADOS")
        
        print(Fore.YELLOW + "Escolha o tipo de exportação:")
        print("1. Exportar todos os tombamentos")
        print("2. Exportar dados da sessão atual")
        print("3. Voltar")
        
        choice = input(Fore.CYAN + "\nEscolha uma opção: ").strip()
        
        if choice == "1":
            # Exporta todos os tombamentos
            data = self.performance_optimizer.get_spreadsheet_data_cached()
            export_data = [{"tombamento": t, "sala": s} for t, s in data.items()]
            
            format_choice = input("Formato (excel/csv/json): ").lower().strip()
            if format_choice in ["excel", "csv", "json"]:
                self.enhanced_features.export_results(export_data, format_choice, "todos_tombamentos")
        
        elif choice == "2":
            # Exporta dados da sessão
            session_data = {
                "session_start": self.current_session["start_time"].isoformat(),
                "searches_count": self.current_session["searches_count"],
                "reports_processed": self.current_session["reports_processed"],
                "backups_created": self.current_session["backups_created"]
            }
            
            format_choice = input("Formato (json): ").lower().strip()
            if format_choice == "json":
                self.enhanced_features.export_results([session_data], "json", "sessao")
    
    def handle_compare(self):
        """Manipula comparação de relatórios"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("COMPARAR RELATÓRIOS")
        
        pdf_files = self.model.list_pdf_reports()
        
        if len(pdf_files) < 2:
            EnhancedView.show_message("error", "São necessários pelo menos 2 relatórios para comparação")
            return
        
        print(Fore.YELLOW + "Selecione o primeiro relatório:")
        report1 = EnhancedView.tui_select_pdf(pdf_files)
        
        if not report1:
            return
        
        # Remove o primeiro da lista para o segundo
        remaining_pdfs = [pdf for pdf in pdf_files if pdf != report1]
        print(Fore.YELLOW + "\nSelecione o segundo relatório:")
        report2 = EnhancedView.tui_select_pdf(remaining_pdfs)
        
        if not report2:
            return
        
        report1_path = os.path.join(REPORTS_FOLDER, report1)
        report2_path = os.path.join(REPORTS_FOLDER, report2)
        
        EnhancedView.show_message("loading", "Comparando relatórios")
        result = self.enhanced_features.compare_reports(report1_path, report2_path)
        
        if result and EnhancedView.confirm_dialog("Deseja exportar o resultado da comparação?"):
            export_data = [{
                "comparison_date": datetime.now().isoformat(),
                "report1": report1,
                "report2": report2,
                "common_items": len(result["common"]),
                "unique_to_report1": len(result["only_in_1"]),
                "unique_to_report2": len(result["only_in_2"])
            }]
            self.enhanced_features.export_results(export_data, "json", "comparacao")
    
    def clear_cache(self):
        """Limpa cache do sistema"""
        EnhancedView.clear_terminal()
        EnhancedView.print_header("LIMPAR CACHE")
        
        if EnhancedView.confirm_dialog("Deseja limpar o cache do sistema?"):
            self.performance_optimizer.clear_cache()
            EnhancedView.show_message("success", "Cache limpo com sucesso")
    
    def run_menu(self):
        """Loop principal do menu aprimorado"""
        while True:
            self.view.clear_terminal()
            
            # Status do sistema
            EnhancedView.print_header("CONSULTA DE PATRIMÔNIO - CEDUC", "Sistema Aprimorado")
            print(Fore.CYAN + f"🕐 Sessão iniciada: {self.current_session['start_time'].strftime('%d/%m/%Y %H:%M')}")
            print(Fore.CYAN + f"🔍 Buscas realizadas: {self.current_session['searches_count']}")
            print(Fore.CYAN + f"📄 Relatórios processados: {self.current_session['reports_processed']}")
            print(Fore.CYAN + f"💾 Backups criados: {self.current_session['backups_created']}")
            EnhancedView.print_line()
            
            option = self.view.tui_enhanced_menu()
            
            if option is None or option == "0":
                # Estatísticas finais da sessão
                session_duration = datetime.now() - self.current_session["start_time"]
                print(Fore.MAGENTA + Style.BRIGHT + f"\n📊 Resumo da sessão:")
                print(Fore.CYAN + f"   Duração: {session_duration}")
                print(Fore.CYAN + f"   Buscas: {self.current_session['searches_count']}")
                print(Fore.CYAN + f"   Relatórios: {self.current_session['reports_processed']}")
                print(Fore.CYAN + f"   Backups: {self.current_session['backups_created']}")
                print(Fore.MAGENTA + Style.BRIGHT + "\nPrograma encerrado. Até mais 👋")
                break
            
            try:
                if option in ["1", "2", "3", "4"]:
                    self.handle_search(option)
                elif option == "5":
                    self.handle_pdf_verification()
                elif option == "6":
                    self.handle_statistics()
                elif option == "7":
                    self.handle_backup()
                elif option == "8":
                    self.handle_restore()
                elif option == "9":
                    self.handle_configuration()
                elif option == "10":
                    self.handle_export()
                elif option == "11":
                    self.handle_compare()
                elif option == "12":
                    self.clear_cache()
                    
            except Exception as e:
                EnhancedView.show_message("error", f"Erro na operação: {str(e)}")
            
            input(EnhancedView.highlight_key(
                "\nPressione ENTER para continuar...",
                "ENTER",
                Fore.GREEN
            ))
    
    def handle_search(self, option):
        """Manipula busca de itens"""
        criterion_name, column_index = SEARCH_COLUMNS[option]
        value = input(Fore.YELLOW + f"Digite o valor para {criterion_name.upper()}: ")
        
        self.view.clear_terminal()
        self.view.print_header(f"RESULTADO DA BUSCA - {criterion_name.upper()}")
        self.search_items(column_index, value, criterion_name)
    
    def handle_pdf_verification(self):
        """Manipula verificação de PDF"""
        self.view.clear_terminal()
        self.view.print_header("VERIFICAÇÃO DE ITENS DO RELATÓRIO PDF")
        
        pdf_files = self.model.list_pdf_reports()
        
        if not pdf_files:
            input(self.view.highlight_key(
                "\nPressione ENTER para voltar ao menu...",
                "ENTER",
                Fore.GREEN
            ))
            return
        
        selected_pdf = self.view.tui_select_pdf(pdf_files)
        
        if selected_pdf is None:
            return
        
        pdf_path = os.path.join(REPORTS_FOLDER, selected_pdf)
        
        self.view.clear_terminal()
        self.view.print_header("RESULTADO DA VERIFICAÇÃO DO RELATÓRIO")
        self.search_items_from_pdf(pdf_path)


# ==================================================
# PONTO DE ENTRADA DA APLICAÇÃO
# ==================================================

if __name__ == "__main__":
    try:
        EnhancedView.show_message("info", "Inicializando sistema de patrimônio...")
        app = EnhancedController()
        app.run_menu()
    except KeyboardInterrupt:
        print(Fore.YELLOW + "\n\nPrograma interrompido pelo usuário")
    except Exception as e:
        print(Fore.RED + f"\n❌ Erro fatal: {e}")
        print(Fore.YELLOW + "Contate o suporte técnico")
        import traceback
        traceback.print_exc()
        input("\nPressione ENTER para sair...")
